
import matplotlib
matplotlib.use("TkAgg")

try:
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
except ImportError:
    from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk as NavigationToolbar2TkAgg
from matplotlib.figure import Figure
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import *
import serial
import matplotlib.animation as animation
from openpyxl import Workbook
import numpy as np



Encoder = []
Sensor_RPM_1 = []
Sensor_RPM_2 = []
Sens1_Fark = []
Sens2_Fark = []
Sens1_Fail=[]
Sens2_Fail=[]


Sens = Workbook()
sensör_verileri = Sens.active
Satır =2





arduinoData = serial.Serial('COM3',9600)


LARGE_FONT= ("GOTHAM", 14)
MEDIUM_FONT = ("GOTHAM",12)
SMALL_FONT = ("GOTHAM",10)
back =("white")
front=("#0F2B51")



f = Figure(figsize=(5,5), dpi=100)
a = f.add_subplot(111)



               '''


sensör_verileri['A1'] = 'ENCODER RPM'
sensör_verileri['B1']= 'SENSÖR1 RPM'
sensör_verileri['C1'] = 'SENSÖR2_RPM'


class RPM(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
     

        tk.Tk.wm_title(self, "Sensör Test Bench ")
        self.geometry("900x450")

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (AnaSayfa , GrafikSayfasi, Tachometre):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")
        
        self.show_frame(AnaSayfa)
      
        
        self.graph = self.frames[GrafikSayfasi]
        self.ani = animation.FuncAnimation(f, self.animate, interval=1000, cache_frame_data=False)


    def show_graph():
            self.show_frame(GrafikSayfasi)
        
    def show_frame(self, cont):
        
        frame = self.frames[cont]
        frame.tkraise()
        
  
      
    
    def animate(self,i):
            
       
        arduinoString = arduinoData.readline().decode('utf-8').rstrip()
        dataArray = arduinoString.split(',')

       
        Encoder.append(dataArray[2])
        Sensor_RPM_1.append(dataArray[0])
        Sensor_RPM_2.append(dataArray[1])
        enc = Encoder[-1]
        sens1 = Sensor_RPM_1[-1]
        sens2 = Sensor_RPM_2[-1]
                
        sens1_fark = float(enc)-float(sens1)
        sens2_fark = float(enc) - float(sens2)
        global Satır   
        sensör_verileri.cell(row = Satır,column = 1).value = enc
        sensör_verileri.cell(row = Satır,column = 2).value = sens1
        sensör_verileri.cell(row = Satır,column = 3).value = sens2
                    
        Satır = Satır+1
        
        if sens1_fark <=5 and float(enc) !=0 :
                Sens1_Fark.append(sens1_fark)
                        
        elif sens1_fark >= 15 and float (enc) !=0:
                Sens1_Fail.append("Basarısız")
                        
        if sens2_fark <= 5 and float(enc) !=0:
                Sens2_Fark.append(sens2_fark)
                        
        elif sens2_fark >= 15 and float (enc) !=0:
                Sens2_Fail.append("Basarısız")
                          
        
        if len(Sens1_Fark)>30 and len(Sens2_Fark) >30:
                self.frames[AnaSayfa].status_test.configure(text="TEST BAŞARIYLA TAMAMLANDI",bg=back,fg=front ) 
          
                sensör_verileri.cell(row= 3,column = 4).value= "Test Sonucu ="
                sensör_verileri.cell(row= 3,column = 5).value ="Başarılı"
                
                
                        
                
        if len(Sens1_Fail)>30 or len (Sens2_Fail) > 30:
                self.frames[AnaSayfa].status_test.configure(text="TEST BAŞARISIZ",bg=back,fg=front ) 
                sensör_verileri.cell(row= 3,column = 4).value= "Test Sonucu ="
                sensör_verileri.cell(row= 3,column = 5).value ="Başarısız"
                
                                
                

      
        a.plot(Sensor_RPM_1, 'b', Sensor_RPM_2, 'g', Encoder, 'r')
        a.set_ylabel("RPM Degerleri")
        a.set_xlabel("Zaman")
        a.legend(["Sens1","Sens2","Encoder"], loc ="lower right")
        
        self.frames[AnaSayfa].encoder_label.configure(text="Encoder RPM = {}".format(enc))
        self.frames[AnaSayfa].sensor1_label.configure(text="Sensör1 RPM = {}".format(sens1))
        self.frames[AnaSayfa].sensor2_label.configure(text="Sensör2 RPM = {}".format(sens2)) 
                   
        
        
  
        
class AnaSayfa(tk.Frame):

    def __init__(self,parent,controller):
        tk.Frame.__init__(self,parent, bg=back)
        self.controller = controller
        
        self.frame1 = tk.Frame(self, bg ='white' )
        self.frame1.pack(side='left' )
        
       
        self.frame2 = tk.Frame(self,bg = back)
        self.frame2.pack(side='left')
        
        self.frame3 = tk.Frame(self,bg = back)
        self.frame3.pack(side='left', expand ='true', fill='y')
    

        def Start():
          
             self.status_label.configure(text="Test Başlatıldı", bg=back,fg=front )

             
        def Stop():
     
             self.status_label.configure(text="Test Durduruldu. ", bg=back,fg=front) 
             
        
        label = tk.Label(self.frame2, text="HIZ SENSÖRÜ TESTİ" , font=LARGE_FONT,bg=back,fg=front)
        label.pack(ipadx= 10,pady=10, side='top')
        
        
        
        button4 = ttk.Button(self.frame2,  text ='Testi Başlat', command=Start)
                            
        button4.pack(ipadx=10 ,pady=10, padx=10 )
           
        
        button4 = ttk.Button(self.frame2, text="Testi Durdur", command=Stop)
        button4.pack(ipadx=10, pady=10, padx=10)
        

        self.encoder_label = tk.Label(self.frame2, text="Encoder RPM = ", font=MEDIUM_FONT,bg=back)
        self.encoder_label.pack(pady=10,padx=10)

        self.sensor1_label = tk.Label(self.frame2, text="Sensör1 RPM = ", font=MEDIUM_FONT ,bg=back)
        self.sensor1_label.pack(pady=10,padx=10)

        self.sensor2_label = tk.Label(self.frame2, text="Sensör2 RPM = ", font=MEDIUM_FONT ,bg=back)
        self.sensor2_label.pack(pady=10,padx=10) 
        

        
 
              
        def Excel():
               dosya_adi = filedialog.asksaveasfilename(defaultextension = ".xlsx", filetypes=[("Excel Dosyası","*.xlsx")])
          
               if dosya_adi:
                     Sens.save(dosya_adi)

                        
                                        

           
        self.adsoyad_label= tk.Label(self.frame3, text="    ", font=SMALL_FONT,bg=back,fg=front)
        self.adsoyad_label.pack(padx=10)
                        
        self.adsoyad_label= tk.Label(self.frame3, text="Ad Soyad ", font=SMALL_FONT,bg=back,fg=front)
        self.adsoyad_label.pack(padx=10)
        self.ad_entry = tk.Entry(self.frame3)
        self.ad_entry.pack(padx=10 )
        
        self.serinumara_label= tk.Label(self.frame3, text="Sensör Seri No: ",  font=SMALL_FONT,bg=back,fg=front)
        self.serinumara_label.pack(padx=10,)
        self.serinumara_entry = tk.Entry(self.frame3)
        self.serinumara_entry.pack(padx=10)
 
        
        
        def submit():
             ad=self.ad_entry.get()
             seri_numara=self.serinumara_entry.get()
                 
               
             sensör_verileri.cell(row= 1,column = 4).value = "Testi Yapan Kişi:  "
             sensör_verileri.cell(row = 2,column = 4).value = "Sensör Seri Numarası: " 
             
             sensör_verileri.cell(row= 1,column = 5).value = ad
             sensör_verileri.cell(row = 2,column = 5).value = seri_numara
             
             self.ad_entry.delete(0,END)
             self.serinumara_entry.delete(0,END)
             
              
             
                
        submit_button =ttk.Button(self.frame3, text= "Kaydet",  command =submit)
        submit_button.pack(padx= 10, pady= 10)  
        self.adsoyad_label= tk.Label(self.frame3, text="    ", font=SMALL_FONT,bg=back,fg=front)
        self.adsoyad_label.pack(padx=10)
                
        
        
        
        button3 = ttk.Button(self.frame3, text="Grafiği Görüntüle", command=lambda : controller.show_frame(GrafikSayfasi))
        button3.pack(ipadx=19 ,pady=10, padx=10)
        

         
        button4 = ttk.Button(self.frame3, text="Testi Excel'e Kaydet", command = Excel)
        button4.pack(ipadx=10 ,pady=10, padx=10)
   
        
        
        
        def DonanımTesti():

                self.status_label.configure(text="Donanım Test Edildi", bg=back,fg=front )
              
                
        button3 = ttk.Button(self.frame3, text="Donanımı Test Et",  command= DonanımTesti)
        button3.pack(ipadx=20 ,pady=10, padx=10)
        
        
        
        button3 = ttk.Button(self.frame3, text="Tachometre Testine Geç",
                            command=lambda : controller.show_frame(Tachometre))
        button3.pack(ipadx=15 ,pady=10, padx=10, side = 'bottom')
        
        
        self.status_label = tk.Label(self.frame2, text="", font=MEDIUM_FONT, bg=back)
        self.status_label.pack(pady=10,padx=10)
        
        self.status_test = tk.Label(self.frame2, text="", font=LARGE_FONT, bg=back)
        self.status_test.pack(pady=10,padx=10)
        

class Tachometre(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, bg=back )
        
        
        label = tk.Label(self, text="Test Bench Tachometre Testi", font=LARGE_FONT , bg=  back)
        label.pack(pady=10,padx=10)

        button1 = ttk.Button(self, text=" Sensör Hız Testine Dön",
                            command=lambda: controller.show_frame(AnaSayfa))
        button1.pack() 
      
          


class GrafikSayfasi(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent , bg=back )
        
        
        label = tk.Label(self, text="Sensör ve Encoder RPM Grafiği!", font=LARGE_FONT, bg =back)
        label.pack(padx= 10, pady=10)

        button1 = ttk.Button(self, text="Ana Sayfaya Dön", command=lambda: controller.show_frame(AnaSayfa) )
        button1.pack()


        canvas = FigureCanvasTkAgg(f, self)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)

        toolbar = NavigationToolbar2TkAgg(canvas, self)
        toolbar.update()
        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)


app = RPM()
app.mainloop()


import matplotlib
from matplotlib import pyplot 
matplotlib.use("TkAgg")

try:
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
except ImportError:
    from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk as NavigationToolbar2TkAgg
from matplotlib.figure import Figure
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import *
from PIL import Image, ImageTk
import matplotlib.animation as animation
from gpiozero import Button
from gpiozero import LED
import RPi.GPIO as GPIO
from openpyxl import Workbook
from time import sleep
import serial



Tachometre_dizi1 = []
Tachometre_dizi2 = []
Encoder = []
Sensor_RPM_1 = []
Sensor_RPM_2 = []
Sens1_Fark = []
Sens2_Fark = []
Sens1_Fail=[]
Sens2_Fail=[]
Phase_Diff= []




Sens = Workbook()
sensör_verileri = Sens.active
Satır =2




fig = Figure(figsize=(5, 4), dpi=100)
ax = fig.add_subplot(111)


led_anasayfa =LED(6)
led_success =LED(2)
led_continue = LED(17)
led_fail= LED(4)





arduinoData = serial.Serial('/dev/ttyUSB0', 9600)


LARGE_FONT= ("GOTHAM", 14)
MEDIUM_FONT = ("GOTHAM",12)
SMALL_FONT = ("GOTHAM",10)
back =("white")
front=("#0F2B51")




Acıklama = ''' Hız Sensörü Test Bench projesi, TMS Ar-Ge tarafından geliştirilmiş projedir.
Projenin amacı hızlı trenlerde arıza veren hız sensörlerinin testini gerçekleştirmektir.
Encoder RPM Motor hızı için referans alacağımız değerdir. Test Başlat butonuna basıldığında motor çalışmaya başlar. 
Sensörlerden ve Encoder'den veri okumaya başlarız. Ekrana yazdırılan değerlerden RPM değerlerini takip edebilir 
ya da grafiği göster butonuna tıklayarak yine Sensör RPM değerlerini gözlemleyebilirsiniz. 
Test yeterli veri setini topladığı zaman ekrana "Test Başarıyla tamamlandı" yazısı gösterecektir.
Test Başarısız ise "Test Başarısız" yazısını göreceksiniz. 

               '''

sensör_verileri['A1'] = 'ENCODER RPM'
sensör_verileri['B1']= 'SENSÖR1 RPM'
sensör_verileri['C1'] = 'SENSÖR2_RPM'


class RPM(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        
        tk.Tk.__init__(self, *args, **kwargs)
     
        tk.Tk.wm_title(self, "Sensör Test Bench ")
        self.geometry("900x450")

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (AnaSayfa , GrafikSayfasi, Tachometre):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")
        
        self.show_frame(AnaSayfa)
      
        
        self.graph = self.frames[GrafikSayfasi]
        self.ani = animation.FuncAnimation(fig, self.animate, interval=1000, cache_frame_data=False)
        
        self.control = 0
      

    def show_graph():
            
        self.show_frame(GrafikSayfasi)
            
            
    def show_frame(self, cont):
        
        frame = self.frames[cont]
        frame.tkraise()
        
  

       
        

    def animate(self, i):
            arduinoString = arduinoData.readline().decode('utf-8').rstrip()
            dataArray = arduinoString.split(',')
        

            global Satır
            global led_continue
            global led_success
            global led_fail
            
            
            Phase_Diff.append(float(dataArray[5]))
            Tachometre_dizi2.append(float(dataArray[4]))    
            Tachometre_dizi1.append(float(dataArray[3]))   
            Encoder.append(float(dataArray[2]))
            Sensor_RPM_1.append(float(dataArray[0]))
            Sensor_RPM_2.append(float(dataArray[1]))
            tach1 = Tachometre_dizi1[-1]
            tach2 = Tachometre_dizi2[-1]  
            enc = Encoder[-1]
            sens1 = Sensor_RPM_1[-1]
            sens2 = Sensor_RPM_2[-1]
            phase_diff= Phase_Diff[-1]
         

            sens1_fark = (enc) - (sens1)
            sens2_fark = (enc) - (sens2)
          

            sensör_verileri.cell(row=Satır, column=1).value = enc
            sensör_verileri.cell(row=Satır, column=2).value = sens1
            sensör_verileri.cell(row=Satır, column=3).value = sens2

            Satır = Satır + 1

            if sens1_fark < 10 and (enc) != 0:
                Sens1_Fark.append(sens1_fark)
            elif sens1_fark >= 10 and (enc) != 0:
                Sens1_Fail.append("Basarısız")

            if sens2_fark < 10 and (enc) != 0:
                Sens2_Fark.append(sens2_fark)
            elif sens2_fark >= 10 and (enc) != 0:
                Sens2_Fail.append("Basarısız")

            if len(Sens1_Fark) > 30 and len(Sens2_Fark) > 30:
                self.frames[AnaSayfa].status_test.configure(text="TEST BAŞARIYLA TAMAMLANDI", bg=back, fg=front)
                led_continue.off()
                led_success.on()
                led_fail.off()
                sensör_verileri.cell(row=3, column=4).value = "Test Sonucu ="
                sensör_verileri.cell(row=3, column=5).value = "Başarılı"

            if len(Sens1_Fail) > 30 or len(Sens2_Fail) > 30:
                self.frames[AnaSayfa].status_test.configure(text="TEST BAŞARISIZ", bg=back, fg=front)
                led_continue.off()
                led_fail.on()
                led_success.off()
                sensör_verileri.cell(row=3, column=4).value = "Test Sonucu ="
                sensör_verileri.cell(row=3, column=5).value = "Başarısız"
            
            
            self.frames[AnaSayfa].encoder_label.configure(text="Encoder RPM = {}".format(enc))
            self.frames[AnaSayfa].sensor1_label.configure(text="Sensör1 RPM = {}".format(sens1))
            self.frames[AnaSayfa].sensor2_label.configure(text="Sensör2 RPM = {}".format(sens2))


            self.frames[Tachometre].encoder_label.configure(text="Encoder RPM = {}".format(enc))
            self.frames[Tachometre].tachometre1_label.configure(text="Tachometre1 RPM = {}".format(tach1))
            self.frames[Tachometre].tachometre2_label.configure(text="Tachometre2 RPM = {}".format(tach2))
            self.frames[Tachometre].phase_label.configure(text="Phase Difference = {}".format(phase_diff))           
            
             
            ax.clear()
            ax.plot(Encoder, 'b', Sensor_RPM_1, 'g', Sensor_RPM_2, 'r')
            ax.set_xlabel("ZAMAN", fontsize=10)
            ax.set_ylabel("RPM DEGERLERİ", fontsize=10)
            ax.set_title("Encoder RPM AND Sensör RPM", fontsize=12)
            ax.legend(["Encoder", "Sens1", "Sens2"], loc="upper right")
            
        
  
        
class AnaSayfa(tk.Frame):

    def __init__(self,parent,controller):
        tk.Frame.__init__(self,parent, bg=back)
        self.controller = controller
        
        self.frame1 = tk.Frame(self, bg ='white' )
        self.frame1.pack(side='left' )
        
       
        self.frame2 = tk.Frame(self,bg = back)
        self.frame2.pack(side='left')
        
        self.frame3 = tk.Frame(self,bg = back)
        self.frame3.pack(side='left', expand ='true', fill='y')
        
        self.Test_Stop = LED(18)
        self.Test_Start =LED(27)
        self.led_excel = LED(13)
        self.led_graph = LED(22)
        self.Test_Motor = LED(10)   
        self.led_donanım = LED(26)
        
        global led_continue
        global led_success
        global led_fail
        global led_anasayfa
        global Test_Start
        global Test_Stop
        

       
        
        #External buton konfigürasyonları
        self.button_excel = Button(20, pull_up=False)
        self.buton_donanım_test =Button(21, pull_up =False)
        self.button_graph =Button(16, pull_up=False)
         
        

        def Start():
             self.Test_Start.on()
             sleep(1)
             self.Test_Start.off()
      
             self.status_label.configure(text="Test Başlatıldı", bg=back,fg=front )

             
        def Stop():
             self.Test_Stop.on()  
             sleep(1)
             self.Test_Stop.off() 
             self.status_label.configure(text="Test Durduruldu. ", bg=back,fg=front) 


        
        Tms_logo =Image.open("/home/tmsarge/Desktop/tmssss.jpeg")
        resized_logo = Tms_logo.resize((150,150))
        Tms_logo =ImageTk.PhotoImage(resized_logo)
        logo_label=tk.Label(self.frame1, image =Tms_logo, bg=back)
        logo_label.image = Tms_logo
        logo_label.pack(side='top' ) 


        
        text_box= Text(self.frame1, height=15,width=40)
        text_box.insert("end",Acıklama)
        text_box.config(state='disabled')
        text_box.pack(side='top' )
        
        
             
        
        label = tk.Label(self.frame2, text="HIZ SENSÖRÜ TESTİ" , font=LARGE_FONT,bg=back,fg=front)
        label.pack(ipadx= 10,pady=10, side='top')
        
        
        
        button4 = ttk.Button(self.frame2,  text ='Testi Başlat', command=Start)
                            
        button4.pack(ipadx=10 ,pady=10, padx=10 )
           
        
        button4 = ttk.Button(self.frame2, text="Testi Durdur", command=Stop)
        button4.pack(ipadx=10, pady=10, padx=10)
        

        self.encoder_label = tk.Label(self.frame2, text="Encoder RPM = ", font=MEDIUM_FONT,bg=back)
        self.encoder_label.pack(pady=10,padx=10)

        self.sensor1_label = tk.Label(self.frame2, text="Sensör1 RPM = ", font=MEDIUM_FONT ,bg=back)
        self.sensor1_label.pack(pady=10,padx=10)

        self.sensor2_label = tk.Label(self.frame2, text="Sensör2 RPM = ", font=MEDIUM_FONT ,bg=back)
        self.sensor2_label.pack(pady=10,padx=10) 
        

        
 
              
        def Excel():
               dosya_adi = filedialog.asksaveasfilename(defaultextension = ".xlsx", filetypes=[("Excel Dosyası","*.xlsx")])
               self.led_excel.on()
               sleep(1)
               self.led_excel.off()
               if dosya_adi:
                     Sens.save(dosya_adi)

                        
                                        

           
        self.adsoyad_label= tk.Label(self.frame3, text="    ", font=SMALL_FONT,bg=back,fg=front)
        self.adsoyad_label.pack(padx=10)
                        
        self.adsoyad_label= tk.Label(self.frame3, text="Ad Soyad ", font=SMALL_FONT,bg=back,fg=front)
        self.adsoyad_label.pack(padx=10)
        self.ad_entry = tk.Entry(self.frame3)
        self.ad_entry.pack(padx=10 )
        
        self.serinumara_label= tk.Label(self.frame3, text="Sensör Seri No: ",  font=SMALL_FONT,bg=back,fg=front)
        self.serinumara_label.pack(padx=10,)
        self.serinumara_entry = tk.Entry(self.frame3)
        self.serinumara_entry.pack(padx=10)
 
        
        
        def submit():
             ad=self.ad_entry.get()
             seri_numara=self.serinumara_entry.get()
                 
               
             sensör_verileri.cell(row= 1,column = 4).value = "Testi Yapan Kişi:  "
             sensör_verileri.cell(row = 2,column = 4).value = "Sensör Seri Numarası: " 
             
             sensör_verileri.cell(row= 1,column = 5).value = ad
             sensör_verileri.cell(row = 2,column = 5).value = seri_numara
             
             self.ad_entry.delete(0,END)
             self.serinumara_entry.delete(0,END)
             
              
             
                
        submit_button =ttk.Button(self.frame3, text= "Kaydet",  command =submit)
        submit_button.pack(padx= 10, pady= 10)  
        self.adsoyad_label= tk.Label(self.frame3, text="    ", font=SMALL_FONT,bg=back,fg=front)
        self.adsoyad_label.pack(padx=10)
                
        
        
        
        button3 = ttk.Button(self.frame3, text="Grafiği Görüntüle", command=lambda : (controller.show_frame(GrafikSayfasi), self.led_graph.on(), sleep(2), self.led_graph.off()))
        button3.pack(ipadx=19 ,pady=10, padx=10)
        
        self.button_graph.when_pressed = lambda: (controller.show_frame(GrafikSayfasi) , self.led_graph.on(), sleep(2), self.led_graph.off())      
        
        
         
        button4 = ttk.Button(self.frame3, text="Testi Excel'e Kaydet", command = Excel)
        button4.pack(ipadx=10 ,pady=10, padx=10)
        
        self.button_excel.when_pressed = Excel
        
        
        
        
        def DonanımTesti():
                led_continue.off()
                self.Test_Motor.on()
                sleep(1)
                self.Test_Motor.off()
                sleep(0.5)
                led_success.on()
                sleep(0.5)
                led_continue.on()
                sleep(0.5)
                led_fail.on()
                sleep(0.5)
                self.led_graph.on()
                sleep(0.5)
                led_anasayfa.on()
                sleep(0.5)
                self.led_excel.on()
                sleep(0.5)
                self.led_donanım.on()
                sleep(0.5)
                led_continue.off()
                led_success.off()
                led_fail.off()
                self.Test_Motor.off()
                self.led_graph.off()
                self.led_excel.off()
                led_anasayfa.off()
                self.led_donanım.off()
                self.status_label.configure(text="Donanım Test Edildi", bg=back,fg=front )
              
                
        button3 = ttk.Button(self.frame3, text="Donanımı Test Et",  command= DonanımTesti)
        button3.pack(ipadx=20 ,pady=10, padx=10)
        
        self.buton_donanım_test.when_pressed = DonanımTesti
        
                 
        
        button3 = ttk.Button(self.frame3, text="Tachometre Testine Geç",
                            command= lambda:controller.show_frame(Tachometre))
        button3.pack(ipadx=15 ,pady=10, padx=10, side = 'bottom')
        
        
        self.status_label = tk.Label(self.frame2, text="", font=MEDIUM_FONT, bg=back)
        self.status_label.pack(pady=10,padx=10)
        
        self.status_test = tk.Label(self.frame2, text="", font=LARGE_FONT, bg=back)
        self.status_test.pack(pady=10,padx=10)
        

class Tachometre(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent, bg=back )
                
        self.frame1 = tk.Frame(self, bg = back)
        self.frame1.pack(side='left', expand ='true', fill='y' )
        
        self.label = tk.Label(self.frame1, text="Test Bench Tachometre Testi", font=LARGE_FONT , bg=  back)
        self.label.pack(pady=10,padx=10)
        
                            
        self.delik_sayisi= tk.Label(self.frame1, text="Delik Sayısı ", font=SMALL_FONT,bg=back,fg=front)
        self.delik_sayisi.pack(padx=10)
        self.sayi_entry = tk.Entry(self.frame1)
        self.sayi_entry.pack(padx=10 )
      
        
        def ilet():
                delik_sayi =self.sayi_entry.get()
                self.sayi_entry.delete(0,END)
                
        
        kaydet_button=ttk.Button(self.frame1, text= "İlet",  command =ilet)
        kaydet_button.pack(padx= 10, pady= 10) 
        
   
        
        
        self.encoder_label = tk.Label(self.frame1, text="Encoder RPM = ", font=MEDIUM_FONT,bg=back)
        self.encoder_label.pack(pady=10,padx=10)
        
        self.tachometre1_label= tk.Label(self.frame1, text="Tachometre1 RPM = ", font=MEDIUM_FONT,bg=back)
        self.tachometre1_label.pack(pady=10,padx=10)
        
        self.tachometre2_label= tk.Label(self.frame1, text="Tachometre2 RPM = ", font=MEDIUM_FONT,bg=back)
        self.tachometre2_label.pack(pady=10,padx=10)
        
        self.phase_label= tk.Label(self.frame1, text="Phase Difference= ", font=MEDIUM_FONT,bg=back)
        self.phase_label.pack(pady=10,padx=10)
        
        
        
        
        
        button9 = ttk.Button(self.frame1, text=" Sensör Hız Testine Dön",
                                 command = lambda:(controller.show_frame(AnaSayfa)))
        button9.pack(ipadx=20 ,pady=10, padx=10) 

        

        
        


class GrafikSayfasi(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent , bg=back )
        
        self.buton_anasayfa =Button(12, pull_up= False)
        
        label = tk.Label(self, text="Sensör ve Encoder RPM Grafiği!", font=LARGE_FONT, bg =back)
        label.pack(padx= 10, pady=10)

        button1 = ttk.Button(self, text="Ana Sayfaya Dön",
                            command=lambda: (controller.show_frame(AnaSayfa) ,led_anasayfa.on(), sleep(2), led_anasayfa.off()))
        button1.pack()
        
               
        
        global led_anasayfa
        self.buton_anasayfa.when_pressed = lambda: (controller.show_frame(AnaSayfa) , led_anasayfa.on(), sleep(2), led_anasayfa.off())      

        canvas = FigureCanvasTkAgg(fig, self)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
       

        toolbar = NavigationToolbar2TkAgg(canvas, self)
        toolbar.update()
        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)


app = RPM()
app.mainloop()
